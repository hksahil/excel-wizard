import streamlit as st
import pandas as pd
from io import BytesIO
import zipfile

st.title('Excel Wizard')

# Function to split an Excel file into separate files for each sheet
import openpyxl
from openpyxl import load_workbook
from io import BytesIO
import zipfile

# Function to split an Excel file into separate files for each sheet with formatting
# Function to split an Excel file into separate files for each sheet with formatting
def split_excel(file):
    # Load the workbook using openpyxl to maintain formatting
    original_wb = load_workbook(file, data_only=False)
    
    output = BytesIO()
    with zipfile.ZipFile(output, 'w') as zf:
        for sheet_name in original_wb.sheetnames:
            # Create a new workbook for each sheet
            new_wb = openpyxl.Workbook()
            new_sheet = new_wb.active
            new_sheet.title = sheet_name

            # Get the original sheet
            original_sheet = original_wb[sheet_name]
            
            # Copy content and formatting from original sheet to new sheet
            for row in original_sheet.iter_rows():
                for cell in row:
                    new_cell = new_sheet[cell.coordinate]
                    new_cell.value = cell.value

                    # Manually copy formatting: font, border, alignment, fill, and number format
                    if cell.has_style:
                        new_cell.font = cell.font.copy()
                        new_cell.border = cell.border.copy()
                        new_cell.alignment = cell.alignment.copy()
                        new_cell.fill = cell.fill.copy()
                        new_cell.number_format = cell.number_format

            # Save the new sheet to an in-memory BytesIO object
            with BytesIO() as sheet_output:
                new_wb.save(sheet_output)
                zf.writestr(f"{sheet_name}.xlsx", sheet_output.getvalue())
    
    output.seek(0)
    return output
# Function to merge (union) multiple Excel files into separate sheets within one Excel file
def merge_excels(files):
    combined_output = BytesIO()
    
    with pd.ExcelWriter(combined_output, engine='xlsxwriter') as writer:
        for i, file in enumerate(files):
            excel_data = pd.ExcelFile(file)
            for sheet_name in excel_data.sheet_names:
                sheet_data = pd.read_excel(file, sheet_name=sheet_name)
                # Use a unique sheet name for each input sheet, e.g., File1_Sheet1, File2_Sheet2, etc.
                new_sheet_name = f"File{i+1}_{sheet_name}"
                sheet_data.to_excel(writer, sheet_name=new_sheet_name, index=False)
    
    combined_output.seek(0)
    
    return combined_output

# File upload options
st.sidebar.title("Excel Wizard Options")
option = st.sidebar.radio("Choose an action", ('Split Excel by Sheets', 'Merge Excel Files'))

# Split Excel File
if option == 'Split Excel by Sheets':
    uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx"])
    if uploaded_file is not None:
        st.write("Processing...")
        split_result = split_excel(uploaded_file)
        st.download_button("Download Split Files (ZIP)", data=split_result, file_name="split_sheets.zip")

# Merge Excel Files
elif option == 'Merge Excel Files':
    uploaded_files = st.file_uploader("Upload multiple Excel files", type=["xlsx"], accept_multiple_files=True)
    if uploaded_files:
        st.write("Processing...")
        merged_result = merge_excels(uploaded_files)
        
        st.download_button("Download Merged File", data=merged_result, file_name="merged_file.xlsx")
